"""
Parser para archivos Excel (.xlsx) de extractos bancarios.
Detecta automáticamente la hoja y la fila de encabezados.
"""

import io
from datetime import date, datetime
from openpyxl import load_workbook

from app.services.parser.base import AbstractParser, ResultadoParseo, TransaccionRaw
from app.services.parser.column_detector import detectar_columnas_llm, MapeoColumnas


class XLSXParser(AbstractParser):
    """
    Parsea extractos bancarios en formato Excel (.xlsx).
    Busca automáticamente la hoja con más datos y la fila de encabezados.
    """

    async def parsear(self, contenido: bytes, nombre_archivo: str) -> ResultadoParseo:
        resultado = ResultadoParseo()

        try:
            wb = load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
        except Exception as e:
            resultado.errores.append(f"No se pudo abrir el archivo Excel: {e}")
            return resultado

        # Seleccionar la hoja con más filas
        hoja = self._seleccionar_hoja(wb)
        if hoja is None:
            resultado.errores.append("No se encontró ninguna hoja con datos.")
            return resultado

        # Leer todas las filas como listas de strings
        todas_las_filas = []
        for fila in hoja.iter_rows(values_only=True):
            todas_las_filas.append([self._celda_a_str(c) for c in fila])

        if len(todas_las_filas) < 2:
            resultado.errores.append("La hoja Excel no tiene suficientes filas.")
            return resultado

        # Encontrar la fila de encabezados (primera fila no vacía)
        idx_header = self._encontrar_header(todas_las_filas)
        headers = todas_las_filas[idx_header]
        filas_datos = todas_las_filas[idx_header + 1:]
        resultado.total_filas = len(filas_datos)

        # Detectar mapeo de columnas con LLM
        mapeo = await detectar_columnas_llm(headers)
        resultado.columnas_detectadas = mapeo.to_dict()

        if not mapeo.es_valido():
            resultado.errores.append(
                f"No se encontraron columnas requeridas. Encabezados detectados: {headers}"
            )
            return resultado

        # Parsear cada fila
        for i, fila in enumerate(filas_datos, start=idx_header + 2):
            if not any(c for c in fila if c):
                resultado.filas_omitidas += 1
                continue

            try:
                tx = self._parsear_fila(fila, mapeo, i)
                resultado.transacciones.append(tx)
            except Exception as e:
                resultado.advertencias.append(f"Fila {i}: {e}")
                resultado.filas_omitidas += 1

        wb.close()
        return resultado

    def _seleccionar_hoja(self, wb):
        """Retorna la hoja activa o la que tenga más filas."""
        if wb.active and wb.active.max_row and wb.active.max_row > 1:
            return wb.active
        # Si la hoja activa está vacía, buscar la más grande
        mejor = None
        max_filas = 0
        for nombre in wb.sheetnames:
            hoja = wb[nombre]
            if hoja.max_row and hoja.max_row > max_filas:
                max_filas = hoja.max_row
                mejor = hoja
        return mejor

    def _encontrar_header(self, filas: list[list[str]]) -> int:
        """
        Encuentra el índice de la fila de encabezados.
        Busca la primera fila con al menos 3 celdas con texto.
        """
        for i, fila in enumerate(filas[:20]):  # Solo buscar en las primeras 20 filas
            celdas_con_texto = sum(1 for c in fila if c and c.strip())
            if celdas_con_texto >= 3:
                return i
        return 0

    def _celda_a_str(self, valor) -> str:
        """Convierte cualquier valor de celda Excel a string."""
        if valor is None:
            return ""
        if isinstance(valor, (datetime, date)):
            return valor.strftime("%Y-%m-%d")
        if isinstance(valor, float):
            # Evitar notación científica
            if valor == int(valor):
                return str(int(valor))
            return str(round(valor, 2))
        return str(valor).strip()

    def _parsear_fila(self, fila: list[str], mapeo: MapeoColumnas, num_fila: int) -> TransaccionRaw:
        def celda(idx: int | None) -> str:
            if idx is None or idx >= len(fila):
                return ""
            return fila[idx].strip()

        descripcion = celda(mapeo.col_descripcion)
        if not descripcion:
            raise ValueError("Descripción vacía")

        fecha = self._normalizar_fecha(celda(mapeo.col_fecha))

        if mapeo.col_cargo is not None and mapeo.col_abono is not None:
            cargo_str = celda(mapeo.col_cargo)
            abono_str = celda(mapeo.col_abono)

            if cargo_str and cargo_str not in ("0", "0.0", "-", ""):
                monto, _ = self._normalizar_monto(cargo_str)
                es_cargo = True
            elif abono_str and abono_str not in ("0", "0.0", "-", ""):
                monto, _ = self._normalizar_monto(abono_str)
                es_cargo = False
            else:
                raise ValueError("Sin monto válido")
        else:
            monto_raw = celda(mapeo.col_monto)
            if not monto_raw:
                raise ValueError("Monto vacío")
            monto, es_cargo = self._normalizar_monto(monto_raw)

        rut = celda(mapeo.col_rut) if mapeo.col_rut is not None else None

        return TransaccionRaw(
            descripcion=descripcion,
            monto=monto,
            es_cargo=es_cargo,
            fecha=fecha,
            rut_comercio=rut or None,
            fila_original=num_fila,
        )
